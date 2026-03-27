import json
import re
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_SOURCE = Path(__file__).resolve().parent.parent / "Cortes 4200.xlsx"
DEFAULT_SHEET = "Ranking 42"
DEFAULT_OUTPUT = Path("trazabilidad-data.json")
SIZES = ["34", "36", "38", "40", "42", "44", "46", "48"]


def parse_int(value) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def normalize_model_code(value) -> str:
    text = str(value or "").strip().upper().replace(" ", "")
    if not text:
        return ""
    m = re.match(r"^(\d{4})(?:-(\d{2}))?$", text)
    if not m:
        return ""
    return m.group(1)


def normalize_sku(value) -> str:
    text = str(value or "").strip().upper().replace(" ", "")
    m = re.match(r"^(\d{4})(?:-(\d{2}))?$", text)
    if not m:
        return text
    model = m.group(1)
    variant = m.group(2) or "00"
    return f"{model}-{variant}"


def load_ex_mapping(workbook) -> dict:
    # Mapa nuevo(42xx) -> ex(41xx/40xx) desde nombres de hojas "Ex ####"
    mapping = {}
    for name in workbook.sheetnames:
        upper = name.upper()
        m_new = re.search(r"\b(42\d{2})\b", upper)
        if not m_new:
            continue
        new_code = m_new.group(1)
        ex_list = re.findall(r"EX\s*(\d{4}(?:-?00)?)", upper)
        if not ex_list:
            continue
        ex_code = normalize_model_code(ex_list[0])
        if ex_code:
            mapping[new_code] = ex_code
    return mapping


def parse_ranking42(source_path: Path, sheet_name: str) -> dict:
    wb = load_workbook(source_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet_name}' en {source_path.name}")

    ex_mapping = load_ex_mapping(wb)
    ws = wb[sheet_name]

    items = []
    for r in range(1, ws.max_row + 1):
        sku_raw = ws.cell(r, 1).value        # A
        fabric = ws.cell(r, 6).value         # F
        size_vals = [ws.cell(r, c).value for c in range(7, 15)]  # G..N
        total_raw = ws.cell(r, 15).value     # O

        sku = normalize_sku(sku_raw)
        model = normalize_model_code(sku_raw)
        if not sku or not model.startswith("42"):
            continue

        total = parse_int(total_raw)
        # Regla negocio: negativos = disponibles
        available_total = abs(total) if total < 0 else 0
        if available_total <= 0:
            continue

        sizes_available = {}
        for size_name, value in zip(SIZES, size_vals):
            n = parse_int(value)
            sizes_available[size_name] = abs(n) if n < 0 else 0

        ex_model = ex_mapping.get(model, "")
        model_display = f"{model}/{ex_model}" if ex_model else model

        items.append({
            "sku": model_display,
            "sku_new": model,
            "sku_new_00": f"{model}-00",
            "sku_ex": ex_model,
            "sku_ex_00": f"{ex_model}-00" if ex_model else "",
            "article": sku,
            "fabric": str(fabric or "").strip(),
            "available_units": available_total,
            "sizes_available": sizes_available,
            "source_row": r,
        })

    items.sort(key=lambda x: x["available_units"], reverse=True)

    return {
        "source_file": str(source_path),
        "sheet": sheet_name,
        "rule": "Negativos en Ranking 42 = disponibles",
        "items_total": len(items),
        "available_units_total": sum(i["available_units"] for i in items),
        "items": items,
    }


def main() -> None:
    if not DEFAULT_SOURCE.exists():
        raise SystemExit(f"No se encontro el archivo: {DEFAULT_SOURCE}")

    payload = parse_ranking42(DEFAULT_SOURCE, DEFAULT_SHEET)
    DEFAULT_OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Trazabilidad exportada a {DEFAULT_OUTPUT.resolve()}")
    print(f"Modelos: {payload['items_total']}")
    print(f"Disponibles: {payload['available_units_total']}")


if __name__ == "__main__":
    main()
