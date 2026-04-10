import json
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_SOURCE = Path(r"C:\Users\manuh\OneDrive - Mohicano Jeans\INVENTARIO 01-04 COMPLETO.xlsx")
DEFAULT_OUTPUT = Path("stock-data.json")
DEFAULT_OUTPUT_CATALOGO_2 = Path("stock-data-catalogo-2.json")
DEFAULT_SHEET = "principal 42"
CATALOGO_2_SHEETS = ("COLE 40", "cole 40", "COLE 41", "cole 41")
SIZE_COLUMNS = {
    "36": 5,
    "38": 6,
    "40": 7,
    "42": 8,
    "44": 9,
    "46": 10,
}


def normalize_article_code(value) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    if not digits:
        return ""

    if len(digits) == 6:
        model = digits[:4]
        variant = digits[4:6]
    elif len(digits) >= 8:
        model = digits[-6:-2]
        variant = digits[-2:]
    else:
        return digits

    return f"{model}-{variant}"


def parse_int(value) -> int:
    if value in (None, ""):
        return 0
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return 0


def build_description(row) -> str:
    tiro = str(row[2] or "").strip()
    bota = str(row[3] or "").strip()
    color = str(row[4] or "").strip()
    parts = [part for part in [tiro, bota, color] if part]
    return " / ".join(parts)


def resolve_sheet_name(workbook, desired_name: str) -> str:
    if desired_name in workbook.sheetnames:
        return desired_name
    desired_normalized = str(desired_name or "").strip().lower()
    for candidate in workbook.sheetnames:
        if str(candidate).strip().lower() == desired_normalized:
            return candidate
    return ""


def parse_sheet_items(ws) -> dict:
    items = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        raw_code = row[1] if len(row) > 1 else None
        sku = normalize_article_code(raw_code)
        if not sku:
            continue

        sizes = {
            size: parse_int(row[col_index]) if len(row) > col_index else 0
            for size, col_index in SIZE_COLUMNS.items()
        }
        total = sum(sizes.values())
        if total <= 0:
            continue

        items[sku] = {
            "article": str(raw_code).strip(),
            "sku": sku,
            "description": build_description(row),
            "sizes": sizes,
            "total": total,
        }
    return items


def merge_stock_items(items_groups) -> dict:
    merged = {}
    for items in items_groups:
        for sku, payload in (items or {}).items():
            if sku not in merged:
                merged[sku] = {
                    "article": payload.get("article", ""),
                    "sku": sku,
                    "description": payload.get("description", ""),
                    "sizes": {size: 0 for size in SIZE_COLUMNS.keys()},
                    "total": 0,
                }
            entry = merged[sku]
            for size in SIZE_COLUMNS.keys():
                entry["sizes"][size] = int(entry["sizes"].get(size, 0)) + int(payload.get("sizes", {}).get(size, 0))
            entry["total"] = sum(entry["sizes"].values())
            if not entry["description"] and payload.get("description"):
                entry["description"] = payload["description"]
            if not entry["article"] and payload.get("article"):
                entry["article"] = payload["article"]
    return merged


def parse_stock_excel(source_path: Path, sheet_name: str = DEFAULT_SHEET) -> dict:
    wb = load_workbook(source_path, read_only=True, data_only=True)
    resolved_sheet = resolve_sheet_name(wb, sheet_name)
    if not resolved_sheet:
        for fallback in ("principal 42", "COLE 42", "Hoja1"):
            resolved_sheet = resolve_sheet_name(wb, fallback)
            if resolved_sheet:
                break
    if not resolved_sheet:
        resolved_sheet = wb.sheetnames[0]
    ws = wb[resolved_sheet]

    items = parse_sheet_items(ws)

    return {
        "source_file": str(source_path),
        "sheet_name": ws.title,
        "items": items,
    }


def parse_stock_excel_catalogo_2(source_path: Path) -> dict:
    wb = load_workbook(source_path, read_only=True, data_only=True)
    used_sheets = []
    parsed_groups = []

    for candidate in CATALOGO_2_SHEETS:
        resolved = resolve_sheet_name(wb, candidate)
        if not resolved or resolved in used_sheets:
            continue
        used_sheets.append(resolved)
        parsed_groups.append(parse_sheet_items(wb[resolved]))

    if not parsed_groups and wb.sheetnames:
        fallback = wb.sheetnames[0]
        used_sheets = [fallback]
        parsed_groups.append(parse_sheet_items(wb[fallback]))

    items = merge_stock_items(parsed_groups)
    return {
        "source_file": str(source_path),
        "sheet_names": used_sheets,
        "items": items,
    }


def main() -> None:
    if not DEFAULT_SOURCE.exists():
        raise SystemExit(f"No se encontro el archivo: {DEFAULT_SOURCE}")

    payload_42 = parse_stock_excel(DEFAULT_SOURCE)
    DEFAULT_OUTPUT.write_text(json.dumps(payload_42, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Stock exportado a {DEFAULT_OUTPUT.resolve()}")
    print(f"Hoja usada: {payload_42['sheet_name']}")
    print(f"SKUs con stock: {len(payload_42['items'])}")

    payload_40_41 = parse_stock_excel_catalogo_2(DEFAULT_SOURCE)
    DEFAULT_OUTPUT_CATALOGO_2.write_text(json.dumps(payload_40_41, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Stock catalogo-2 exportado a {DEFAULT_OUTPUT_CATALOGO_2.resolve()}")
    print(f"Hojas usadas: {', '.join(payload_40_41['sheet_names'])}")
    print(f"SKUs con stock catalogo-2: {len(payload_40_41['items'])}")


if __name__ == "__main__":
    main()
